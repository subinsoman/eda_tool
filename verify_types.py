import openpyxl
import glob
import os

# Find the latest generated report
list_of_files = glob.glob('/home/subinsoman/profile/eda/*_metrics_report_*.xlsx')
latest_file = max(list_of_files, key=os.path.getctime)
print(f"Checking file: {latest_file}")

wb = openpyxl.load_workbook(latest_file)

# 1. Check "Column Metrics" sheet for Data Types
ws_metrics = wb["Column Metrics"]
# Assuming "Data Type" is the 2nd column (Column B)
# Header is in row 1
header_cell = ws_metrics['B1']
if header_cell.value != "Data Type":
    print(f"FAIL: Expected 'Data Type' in B1, found '{header_cell.value}'")
    # Try to find it
    for cell in ws_metrics[1]:
        if cell.value == "Data Type":
            col_idx = cell.column
            print(f"Found 'Data Type' at column index {col_idx}")
            break
else:
    col_idx = 2

allowed_types = {"Numerical", "Categorical", "Date"}
invalid_found = False

for row in ws_metrics.iter_rows(min_row=2, min_col=1, max_col=col_idx, values_only=True):
    col_name = row[0]
    val = row[col_idx-1]
    print(f"Column: {col_name}, Type: {val}")
    if val not in allowed_types:
        print(f"FAIL: Invalid Data Type found: '{val}'")
        invalid_found = True

if not invalid_found:
    print("PASS: All Data Types are valid (Numerical, Categorical, Date)")

# 2. Check "Summary" sheet for Variable Types table
ws_summary = wb["Summary"]
# Find "Variable Types" table
found_table = False
for row in ws_summary.iter_rows(values_only=True):
    if row and row[0] == "Variable Types":
        found_table = True
        continue # Skip header
    
    if found_table:
        if row[0] in ["Numerical", "Categorical", "Date"]:
            print(f"PASS: Found summary row for '{row[0]}': {row[1]}")
        elif row[0] is None or row[0] == "":
            pass # Empty line
        else:
            # Check if we hit the next table or end
            if row[0] in ["Text", "Boolean", "Integer", "Float"]:
                 print(f"FAIL: Found deprecated summary row for '{row[0]}'")

print("Verification Complete")
