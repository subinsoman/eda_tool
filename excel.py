import warnings

# Suppress specific RuntimeWarning from Dask/NumPy
warnings.filterwarnings("ignore", message="invalid value encountered in scalar divide", category=RuntimeWarning)
# Suppress Dask/Pandas FutureWarnings
warnings.filterwarnings("ignore", message="pandas.Int64Index is deprecated", category=FutureWarning)
warnings.filterwarnings("ignore", message="pandas.Float64Index is deprecated", category=FutureWarning)
warnings.filterwarnings("ignore", message="pandas.UInt64Index is deprecated", category=FutureWarning)

import dask.dataframe as dd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from tqdm import tqdm
import scipy.stats as stats
from datetime import datetime
import dask.array as da
import dask
import argparse
from cli_utils import parse_args
from pathlib import Path
import sys
import json

class DataLoader:
    """Responsible for loading data."""
    def __init__(self, input_path: Path, dask_args: dict = None):
        self.input_path = input_path
        self.dask_args = dask_args or {}

    def load_data(self) -> dd.DataFrame:
        print(f"Reading data from: {self.input_path}")
        # Merge default dtype with user-provided args if needed, or just pass kwargs
        # Ensure 'dtype' for Cabin is preserved if not overridden, or just let user override
        default_args = {'dtype': {'Cabin': 'object'}}
        
        # Merge defaults with user args (user args take precedence)
        final_args = default_args.copy()
        final_args.update(self.dask_args)
        
        return dd.read_csv(str(self.input_path), **final_args)

class MetricsCalculator:
    """Responsible for calculating column metrics."""
    def __init__(self, ddf: dd.DataFrame):
        self.ddf = ddf

    def calculate_metrics(self):
        metrics_list = []
        total_rows = len(self.ddf)
        columns = list(self.ddf.columns)
        total_columns = len(columns)
        
        print(f"Processing {total_columns} columns...")
        print("Computing all metrics in parallel (Batch Processing)...")
        
        BATCH_SIZE = 10
        
        with tqdm(total=total_columns, desc="Analyzing columns", unit="col") as pbar:
            for i in range(0, total_columns, BATCH_SIZE):
                batch_cols = columns[i : i + BATCH_SIZE]
                
                batch_tasks = []
                batch_meta = []
                
                # 1. Prepare Batch
                for col in batch_cols:
                    col_data = self.ddf[col]
                    col_type = col_data.dtype
                    
                    # Check if numeric
                    try:
                        is_numeric = np.issubdtype(col_type, np.number)
                    except (TypeError, AttributeError):
                        col_type_str = str(col_type).lower()
                        is_numeric = any(t in col_type_str for t in ['int', 'float', 'number', 'decimal'])
                    
                    # Value-Based Type Inference (Lazy)
                    if not is_numeric and col_type == 'object':
                        try:
                            sample_partition = col_data.partitions[0]
                            
                            # 1. Try Numeric
                            numeric_converted = dd.to_numeric(sample_partition, errors='coerce')
                            valid_numeric_count = numeric_converted.notnull().sum()
                            total_count = sample_partition.shape[0]
                            
                            n_valid, n_total = dask.compute(valid_numeric_count, total_count)
                            
                            if n_total > 0 and n_valid / n_total > 0.9:
                                print(f"  -> Inferred {col} as Numerical")
                                col_data = dd.to_numeric(col_data, errors='coerce')
                                col_type = col_data.dtype
                                is_numeric = True
                            else:
                                # 2. Try Date
                                try:
                                    date_converted = dd.to_datetime(sample_partition, errors='coerce')
                                    valid_date_count = date_converted.notnull().sum()
                                    d_valid = dask.compute(valid_date_count)[0]
                                    
                                    if n_total > 0 and d_valid / n_total > 0.9:
                                        print(f"  -> Inferred {col} as Date")
                                        col_data = dd.to_datetime(col_data, errors='coerce')
                                        col_type = col_data.dtype
                                except:
                                    pass
                        except Exception as e:
                            print(f"  -> Type inference failed for {col}: {e}")
                    
                    # Low Cardinality Check
                    try:
                        unique_sample = col_data.unique().head(11, npartitions=-1, compute=True)
                        is_low_cardinality = len(unique_sample) <= 10
                    except:
                        is_low_cardinality = False

                    # Build Lazy Computations
                    lazy_computations = {
                        'null_count': col_data.isnull().sum(),
                        'non_null_count': col_data.count(),
                        'unique_count': col_data.nunique(),
                        'memory_size': col_data.memory_usage(deep=True)
                    }
                    
                    if is_numeric:
                        lazy_computations.update({
                            'mean': col_data.mean(),
                            'std': col_data.std(),
                            'min': col_data.min(),
                            'max': col_data.max(),
                            'q1': col_data.quantile(0.25),
                            'median': col_data.quantile(0.5),
                            'q3': col_data.quantile(0.75),
                            'q10': col_data.quantile(0.10),
                            'q90': col_data.quantile(0.90),
                            'var': col_data.var(),
                            'zero_count': (col_data == 0).sum(),
                            'negative_count': (col_data < 0).sum(),
                            'infinite_count': ((col_data == np.inf) | (col_data == -np.inf)).sum(),
                            'skew': ((col_data - col_data.mean()) ** 3).mean() / (col_data.std() ** 3),
                            'kurt': ((col_data - col_data.mean()) ** 4).mean() / (col_data.std() ** 4) - 3,
                            'p70': col_data.quantile(0.70),
                            'mad': (col_data - col_data.mean()).abs().mean(),
                            'median_ad': (col_data - col_data.quantile(0.5)).abs().quantile(0.5),
                        })
                        
                    lazy_computations.update({
                        'value_counts': col_data.value_counts().head(20),
                        'empty_string': (col_data == '').sum() if col_data.dtype == 'object' else None,
                        'distinct_count': col_data.nunique()
                    })
                    
                    if col_data.dtype == 'object':
                        str_lens = col_data.str.len()
                        lazy_computations.update({
                            'text_len_mean': str_lens.mean(),
                            'text_len_min': str_lens.min(),
                            'text_len_max': str_lens.max()
                        })
                    
                    batch_tasks.append(lazy_computations)
                    batch_meta.append({
                        'col': col,
                        'is_numeric': is_numeric,
                        'is_low_cardinality': is_low_cardinality,
                        'col_type': col_type
                    })

                # 2. Compute Batch
                if batch_tasks:
                    batch_results = dask.compute(batch_tasks)[0]
                    
                    # 3. Process Results
                    for computed, meta in zip(batch_results, batch_meta):
                        col = meta['col']
                        is_numeric = meta['is_numeric']
                        is_low_cardinality = meta['is_low_cardinality']
                        col_type = meta['col_type']
                        
                        null_count = computed['null_count']
                        non_null_count = computed['non_null_count']
                        unique_count = computed['unique_count']
                        
                        col_type_str = str(col_type).lower()
                        
                        # Determine user-friendly data type
                        if is_numeric and 'bool' not in col_type_str:
                            friendly_type = 'Numerical'
                        elif 'datetime' in col_type_str:
                            friendly_type = 'Date'
                        else:
                            friendly_type = 'Categorical'

                        metrics = {
                            'Column Name': col,
                            'Data Type': friendly_type,
                            'Raw Data Type': str(col_type),
                            'Null Count': null_count,
                            'Null Percentage': (null_count / total_rows) if total_rows > 0 else 0,
                            'Unique Values': unique_count,
                            'Cardinality Ratio': (unique_count / non_null_count) if non_null_count > 0 else 0,
                            'Is Potential Key': 'Yes' if unique_count == non_null_count and non_null_count > 0 else 'No',
                            'Is Constant': 'Yes' if unique_count == 1 else 'No',
                            'Type Consistency': 1.0
                        }
                        
                        if is_numeric:
                            mean_val = computed['mean']
                            std_val = computed['std']
                            min_val = computed['min']
                            max_val = computed['max']
                            q1 = computed['q1']
                            q3 = computed['q3']
                            iqr = q3 - q1
                            median_val = computed['median']
                            
                            zero_count = computed['zero_count']
                            negative_count = computed['negative_count']
                            mad_val = computed['mad']
                            median_ad_val = computed['median_ad']
                            outlier_lower = q1 - 1.5 * iqr
                            outlier_upper = q3 + 1.5 * iqr
                            
                            metrics.update({
                                'Mean': mean_val,
                                'Median': median_val,
                                'Std Dev': std_val,
                                'Variance': computed['var'],
                                'Min': min_val,
                                'Max': max_val,
                                'Range': max_val - min_val,
                                'Mid-Range': (min_val + max_val) / 2,
                                'Q1 (25%)': q1,
                                'Q3 (75%)': q3,
                                'IQR': iqr,
                                'P10 (10%)': computed['q10'],
                                'P70 (70%)': computed['p70'],
                                'P90 (90%)': computed['q90'],
                                'Coefficient of Variation': (std_val / mean_val) if mean_val != 0 else 0,
                                'Mean Absolute Deviation': mad_val,
                                'Median Absolute Deviation': median_ad_val,
                                'Range Ratio': max_val / min_val if min_val != 0 else np.inf,
                                'Zero Count': zero_count,
                                'Zero Percentage': (zero_count / non_null_count) if non_null_count > 0 else 0,
                                'Negative Count': negative_count,
                                'Negative Percentage': (negative_count / non_null_count) if non_null_count > 0 else 0,
                                'Outlier Lower Bound': outlier_lower,
                                'Outlier Upper Bound': outlier_upper,
                                'Skewness': computed['skew'],
                                'Kurtosis': computed['kurt'],
                                'Excess Kurtosis': computed['kurt'] - 3,
                                'Infinite Count': computed['infinite_count'],
                                'Infinite Percentage': (computed['infinite_count'] / non_null_count) if non_null_count > 0 else 0,
                            })
                        
                        try:
                            value_counts = computed['value_counts']
                            metrics['chart_data'] = {
                                'type': 'categorical',
                                'labels': [str(x)[:20] for x in value_counts.index.tolist()],
                                'values': value_counts.values.tolist()
                            }

                            if not is_numeric or is_low_cardinality:
                                mode_val = value_counts.index[0] if len(value_counts) > 0 else None
                                mode_freq = value_counts.iloc[0] if len(value_counts) > 0 else 0
                                probs = value_counts / value_counts.sum()
                                entropy_val = stats.entropy(probs) if len(probs) > 0 else 0

                                top_5_dict = {}
                                for i, (val, count) in enumerate(list(value_counts.head(5).items())):
                                    top_5_dict[str(val)[:30]] = int(count)
                                
                                metrics.update({
                                    'Mode': str(mode_val)[:50],
                                    'Mode Frequency': mode_freq,
                                    'Mode Percentage': (mode_freq / non_null_count) if non_null_count > 0 else 0,
                                    'Dominant Category': str(mode_val)[:50],
                                    'Dominant Category Percentage': (mode_freq / non_null_count) if non_null_count > 0 else 0,
                                    'Top 5 Values': str(top_5_dict)[:100],
                                    'Concentration (Top 10)': (value_counts.sum() / non_null_count) if non_null_count > 0 else 0,
                                    'Entropy (Top 10)': entropy_val,
                                    'Distinct Ratio': (unique_count / total_rows) if total_rows > 0 else 0,
                                })
                                
                                if computed.get('empty_string') is not None:
                                    metrics['Empty String Count'] = computed['empty_string']
                                
                                if computed.get('text_len_mean') is not None:
                                    metrics.update({
                                        'Text Length Mean': computed['text_len_mean'],
                                        'Text Length Min': computed['text_len_min'],
                                        'Text Length Max': computed['text_len_max']
                                    })
                                    
                        except Exception as e:
                            if not is_numeric:
                                metrics.update({
                                    'Mode': 'Error',
                                    'Mode Frequency': 0,
                                    'Error': str(e)[:50]
                                })
                        
                        duplicate_count = non_null_count - unique_count
                        metrics['Duplicate Count'] = duplicate_count
                        metrics['Duplicate Percentage'] = (duplicate_count / non_null_count) if non_null_count > 0 else 0
                        metrics['Memory Size'] = computed['memory_size']
                        
                        metrics_list.append(metrics)
                        pbar.update(1)
        
        return metrics_list, total_rows

class ReportGenerator:
    """Responsible for generating the Excel report."""
    def __init__(self, metrics_list, total_rows, output_path):
        self.metrics_list = metrics_list
        self.total_rows = total_rows
        self.output_path = output_path
        self.wb = Workbook()
        
        # Convert metrics list to dict for easier processing
        self.metrics_dict = {}
        for metric in self.metrics_list:
            for key, value in metric.items():
                if key not in self.metrics_dict:
                    self.metrics_dict[key] = []
                self.metrics_dict[key].append(value)

    def generate_report(self):
        print("\nGenerating Excel report...")
        
        # Create sheets
        self._create_summary_sheet()
        self._create_definitions_sheet()
        self._create_metrics_sheet()
        self._create_analysis_sheet()
        self._create_chart_data_sheet()
        
        # Set zoom level
        for sheet in self.wb.worksheets:
            sheet.sheet_view.zoomScale = 85
            
        self.wb.save(self.output_path)
        print(f"\n✓ Excel report saved as '{self.output_path}'")

    def _create_summary_sheet(self):
        ws_summary = self.wb.create_sheet("Summary", 0)
        ws_summary.sheet_view.showGridLines = False
        
        # Dashboard Title
        ws_summary.merge_cells('B2:H3')
        title_cell = ws_summary['B2']
        title_cell.value = "Data Quality & Insights Dashboard"
        title_cell.font = Font(bold=True, size=24, color="2F5597")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Timestamp
        ws_summary.merge_cells('I2:K3')
        time_cell = ws_summary['I2']
        time_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        time_cell.font = Font(size=10, color="7F7F7F", italic=True)
        time_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Calculate summary statistics
        null_counts = [m for m in self.metrics_dict['Null Count'] if isinstance(m, (int, float))]
        data_types = self.metrics_dict['Data Type']
        
        numerical_count = sum(1 for dt in data_types if dt == 'Numerical')
        categorical_count = sum(1 for dt in data_types if dt == 'Categorical')
        date_count = sum(1 for dt in data_types if dt == 'Date')
        
        total_columns = len(self.metrics_list)
        total_cells = self.total_rows * total_columns
        total_missing_cells = sum(null_counts)
        missing_cells_pct = (total_missing_cells / total_cells) if total_cells > 0 else 0
        
        print("Calculating memory usage from column metrics...")
        total_memory_bytes = sum(m.get('Memory Size', 0) for m in self.metrics_list)
        avg_row_size_bytes = total_memory_bytes / self.total_rows if self.total_rows > 0 else 0
        
        def format_bytes(size):
            power = 2**10
            n = 0
            power_labels = {0 : '', 1: 'KB', 2: 'MB', 3: 'GB', 4: 'TB'}
            while size > power:
                size /= power
                n += 1
            return f"{size:.1f} {power_labels[n]}"
        
        total_memory_str = format_bytes(total_memory_bytes)
        avg_row_size_str = f"{avg_row_size_bytes:.1f} B"
        
        # Insights Generation
        insights = []
        for m in self.metrics_list:
            if m['Unique Values'] == self.total_rows and m['Null Count'] == 0:
                insights.append([f"{m['Column Name']} is uniformly distributed", "Uniform"])
            if m['Null Count'] > 0:
                insights.append([f"{m['Column Name']} has {m['Null Count']} ({m['Null Percentage']:.2%}) missing values", "Missing"])
            if 'Skewness' in m and isinstance(m['Skewness'], (int, float)):
                if abs(m['Skewness']) > 1:
                    insights.append([f"{m['Column Name']} is skewed", "Skewed"])
            if m['Data Type'] in ['Categorical'] and m['Unique Values'] > 50:
                 insights.append([f"{m['Column Name']} has a high cardinality: {m['Unique Values']} distinct values", "High Cardinality"])
            if 'Text Length Min' in m and 'Text Length Max' in m:
                if m['Text Length Min'] == m['Text Length Max'] and m['Text Length Min'] > 0:
                     insights.append([f"{m['Column Name']} has constant length {int(m['Text Length Min'])}", "Constant Length"])
            if m['Is Constant'] == 'Yes':
                 insights.append([f"{m['Column Name']} has constant value", "Constant Value"])

        stats_data = [
            ["Dataset Statistics", ""],
            ["Number of Variables", total_columns],
            ["Number of Rows", self.total_rows],
            ["Missing Cells", total_missing_cells],
            ["Missing Cells (%)", f"{missing_cells_pct:.1%}"],
            ["Duplicate Rows", "N/A (Optimization)"],
            ["Duplicate Rows (%)", "N/A"],
            ["Total Size in Memory", total_memory_str],
            ["Average Row Size in Memory", avg_row_size_str]
        ]
        
        types_data = [
            ["Variable Types", ""],
            ["Numerical", numerical_count],
            ["Categorical", categorical_count],
            ["Date", date_count],
        ]
        
        start_row_tables = 5
        self._write_styled_table(ws_summary, stats_data, start_row=start_row_tables, start_col=2, width_key=30, width_val=20)
        self._write_styled_table(ws_summary, types_data, start_row=start_row_tables, start_col=5, width_key=25, width_val=15)
        
        # Insights Table
        if "InsightsData" in self.wb.sheetnames:
            ws_data = self.wb["InsightsData"]
        else:
            ws_data = self.wb.create_sheet("InsightsData")
        ws_data.sheet_state = 'hidden'
        
        for row_idx, row_data in enumerate(insights, 1):
            ws_data.cell(row=row_idx, column=1, value=row_data[0])
            ws_data.cell(row=row_idx, column=2, value=row_data[1])
            
        PAGE_SIZE = 20
        total_insights = len(insights)
        total_pages = (total_insights + PAGE_SIZE - 1) // PAGE_SIZE if total_insights > 0 else 1
        
        control_row = start_row_tables
        ws_summary.cell(row=control_row, column=8, value="Dataset Insights").font = Font(bold=True, size=12, color="FFFFFF")
        ws_summary.cell(row=control_row, column=8).fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        ws_summary.cell(row=control_row, column=8).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_summary.merge_cells(start_row=control_row, start_column=8, end_row=control_row, end_column=9)
        
        ws_summary.cell(row=control_row, column=10, value="Page:").font = Font(color="FFFFFF", bold=True)
        ws_summary.cell(row=control_row, column=10).fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        ws_summary.cell(row=control_row, column=10).alignment = Alignment(horizontal="right", vertical="center")
        
        page_cell = ws_summary.cell(row=control_row, column=11, value=1)
        page_cell.alignment = Alignment(horizontal="center", vertical="center")
        page_cell.font = Font(bold=True)
        page_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        dv = DataValidation(type="whole", operator="between", formula1=1, formula2=total_pages)
        dv.error = 'Invalid Page'
        ws_summary.add_data_validation(dv)
        dv.add(page_cell)
        
        ws_summary.cell(row=control_row, column=12, value=f"of {total_pages}").font = Font(color="FFFFFF")
        ws_summary.cell(row=control_row, column=12).fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        ws_summary.cell(row=control_row, column=12).alignment = Alignment(horizontal="left", vertical="center")
        
        for col in range(8, 13):
            ws_summary.cell(row=control_row, column=col).border = Border(bottom=Side(style='thin', color="BFBFBF"), top=Side(style='thin', color="BFBFBF"))
            
        table_start_row = control_row + 1
        start_col_insights = 8
        
        border_color = "BFBFBF"
        thin_border = Border(left=Side(style='thin', color=border_color), 
                             right=Side(style='thin', color=border_color), 
                             top=Side(style='thin', color=border_color), 
                             bottom=Side(style='thin', color=border_color))
        row_fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for i in range(PAGE_SIZE):
            current_row = table_start_row + i
            row_offset = i + 1
            formula_row_idx = f"($K${control_row}-1)*{PAGE_SIZE} + {row_offset}"
            
            cell_metric = ws_summary.cell(row=current_row, column=start_col_insights)
            cell_metric.value = f'=IF({formula_row_idx} > {total_insights}, "", INDEX(InsightsData!A:A, {formula_row_idx}))'
            cell_metric.font = Font(size=11, color="404040")
            cell_metric.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            cell_metric.border = thin_border
            
            cell_value = ws_summary.cell(row=current_row, column=start_col_insights + 1)
            cell_value.value = f'=IF({formula_row_idx} > {total_insights}, "", INDEX(InsightsData!B:B, {formula_row_idx}))'
            cell_value.font = Font(size=11, color="000000")
            cell_value.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            cell_value.border = thin_border
            
            fill = row_fill_even if i % 2 == 0 else row_fill_odd
            cell_metric.fill = fill
            cell_value.fill = fill
            
        ws_summary.column_dimensions['H'].width = 50
        ws_summary.column_dimensions['I'].width = 30
        ws_summary.column_dimensions['J'].width = 10
        ws_summary.column_dimensions['K'].width = 10
        ws_summary.column_dimensions['L'].width = 10

    def _create_definitions_sheet(self):
        ws_definitions = self.wb.create_sheet("Definitions", 1)
        ws_definitions.sheet_view.showGridLines = False
        ws_definitions.append(['Metric', 'Description'])
        
        definitions = [
            ('Column Name', 'Name of the column in the dataset'),
            ('Data Type', 'Inferred data type of the column'),
            ('Null Count', 'Number of missing (NaN/None) values'),
            ('Null Percentage', 'Percentage of rows that are missing'),
            ('Unique Values', 'Number of distinct values'),
            ('Cardinality Ratio', 'Ratio of unique values to non-null values'),
            ('Is Potential Key', 'Indicates if the column could be a unique identifier'),
            ('Is Constant', 'Indicates if the column has only one unique value'),
            ('Type Consistency', 'Percentage of values matching the inferred type (Placeholder)'),
            ('Mean', 'Average value'),
            ('Median', 'Middle value'),
            ('Std Dev', 'Standard deviation'),
            ('Variance', 'Square of the standard deviation'),
            ('Min', 'Minimum value'),
            ('Max', 'Maximum value'),
            ('Range', 'Difference between Max and Min'),
            ('Mid-Range', '(Min + Max) / 2'),
            ('Q1 (25%)', 'First quartile'),
            ('Q3 (75%)', 'Third quartile'),
            ('IQR', 'Interquartile Range (Q3 - Q1)'),
            ('P10 (10%)', '10th Percentile'),
            ('P70 (70%)', '70th Percentile'),
            ('P90 (90%)', '90th Percentile'),
            ('Coefficient of Variation', 'Ratio of Std Dev to Mean (%)'),
            ('Mean Absolute Deviation', 'Average of absolute deviations from the mean'),
            ('Median Absolute Deviation', 'Median of absolute deviations from the median'),
            ('Range Ratio', 'Max / Min'),
            ('Zero Count', 'Number of zero values'),
            ('Zero Percentage', 'Percentage of zero values'),
            ('Negative Count', 'Number of negative values'),
            ('Negative Percentage', 'Percentage of negative values'),
            ('Outlier Lower Bound', 'Q1 - 1.5 * IQR'),
            ('Outlier Upper Bound', 'Q3 + 1.5 * IQR'),
            ('Skewness', 'Measure of asymmetry'),
            ('Kurtosis', 'Measure of tailedness'),
            ('Excess Kurtosis', 'Kurtosis - 3'),
            ('Infinite Count', 'Number of infinite values'),
            ('Infinite Percentage', 'Percentage of infinite values'),
            ('Memory Size', 'Memory usage of the column'),
            ('Mode', 'Most frequent value'),
            ('Mode Frequency', 'Count of the mode'),
            ('Mode Percentage', 'Percentage of the mode'),
            ('Dominant Category', 'Same as Mode'),
            ('Dominant Category Percentage', 'Same as Mode Percentage'),
            ('Top 5 Values', 'List of top 5 most frequent values'),
            ('Concentration (Top 10)', 'Sum of frequencies of top 10 values / Total'),
            ('Entropy (Top 10)', 'Shannon entropy of the top 10 values'),
            ('Distinct Ratio', 'Unique Values / Total Rows'),
            ('Empty String Count', 'Number of empty strings'),
            ('Text Length Mean', 'Average length of text values'),
            ('Text Length Min', 'Minimum length of text values'),
            ('Text Length Max', 'Maximum length of text values'),
            ('Duplicate Count', 'Number of duplicate values (Non-Null - Unique)'),
            ('Duplicate Percentage', 'Percentage of duplicate values')
        ]
        
        for row in definitions:
            ws_definitions.append(row)
            
        tab_def = Table(displayName="MetricDefinitions", ref=ws_definitions.dimensions)
        style_def = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab_def.tableStyleInfo = style_def
        ws_definitions.add_table(tab_def)
        
        ws_definitions.column_dimensions['A'].width = 30
        ws_definitions.column_dimensions['B'].width = 80
        for cell in ws_definitions['B']:
            cell.alignment = Alignment(wrap_text=True)

    def _create_metrics_sheet(self):
        if 'Sheet' in self.wb.sheetnames:
            ws = self.wb['Sheet']
        else:
            ws = self.wb.create_sheet("Column Metrics")
        ws.title = "Column Metrics"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        columns = [k for k in self.metrics_dict.keys() if k != 'chart_data']
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        for row_idx, metric_data in enumerate(self.metrics_list):
            for col_idx, col_name in enumerate(columns, 1):
                value = metric_data.get(col_name)
                cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                if isinstance(value, (int, float)) and not np.isnan(value):
                    if col_idx > 2:
                        if 'Percentage' in col_name or 'Ratio' in col_name or 'Consistency' in col_name or 'Concentration' in col_name or 'Variation' in col_name:
                             cell.number_format = '0.00%'
                        elif abs(value) < 1:
                            cell.number_format = '0.0000'
                        elif abs(value) < 1000:
                            cell.number_format = '0.00'
                        else:
                            cell.number_format = '#,##0.00'
                
                if col_name == 'Null Percentage' and isinstance(value, (int, float)) and value > 0.5:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif col_name == 'Is Potential Key' and value == 'Yes':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif col_name == 'Cardinality Ratio' and isinstance(value, (int, float)) and value > 0.95:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
        ws.freeze_panes = "A2"

    def _create_analysis_sheet(self):
        ws_analysis = self.wb.create_sheet("Column Analysis", 1)
        ws_analysis.sheet_view.showGridLines = False
        
        ws_analysis.merge_cells('B2:E3')
        title_cell = ws_analysis['B2']
        title_cell.value = "Individual Column Analysis"
        title_cell.font = Font(bold=True, size=20, color="2F5597")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        ws_analysis['B5'] = "Select Column:"
        ws_analysis['B5'].font = Font(bold=True, size=12)
        ws_analysis['B5'].alignment = Alignment(horizontal="right", vertical="center")
        
        selector_cell = ws_analysis['C5']
        first_col_name = self.metrics_list[0]['Column Name'] if self.metrics_list else ""
        selector_cell.value = first_col_name
        selector_cell.font = Font(size=12)
        selector_cell.border = Border(bottom=Side(style='thin'))
        selector_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        num_rows = len(self.metrics_list)
        validation_range = f"'Column Metrics'!$A$2:$A${num_rows + 1}"
        dv = DataValidation(type="list", formula1=validation_range, allow_blank=True)
        ws_analysis.add_data_validation(dv)
        dv.add(selector_cell)
        
        general_stats = [
            ("Data Type", "Data Type", "@"),
            ("Raw Data Type", "Raw Data Type", "@"),
            ("Unique Values", "Unique Values", "#,##0"),
            ("Cardinality Ratio", "Cardinality Ratio", "0.0%"),
            ("Null Count", "Null Count", "#,##0"),
            ("Null Percentage", "Null Percentage", "0.0%"),
            ("Infinite Count", "Infinite Count", "#,##0"),
            ("Infinite Percentage", "Infinite Percentage", "0.0%"),
            ("Memory Size", "Memory Size", "#,##0")
        ]
        
        numeric_stats = [
            ("Mean", "Mean", "#,##0.00"),
            ("Min", "Min", "#,##0.00"),
            ("Max", "Max", "#,##0.00"),
            ("Zero Count", "Zero Count", "#,##0"),
            ("Zero Percentage", "Zero Percentage", "0.0%"),
            ("Negative Count", "Negative Count", "#,##0"),
            ("Negative Percentage", "Negative Percentage", "0.0%")
        ]
        
        start_row = 8
        self._create_styled_metric_table(ws_analysis, "General Statistics", general_stats, start_row, 2)
        self._create_styled_metric_table(ws_analysis, "Numeric Statistics", numeric_stats, start_row, 5)
        
        # Chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Value vs Frequency"
        chart.y_axis.title = "Frequency"
        chart.x_axis.title = "Value"
        chart.height = 10
        chart.width = 20
        
        if "ChartData" not in self.wb.sheetnames:
             self.wb.create_sheet("ChartData")
        ws_chart_data = self.wb["ChartData"]
        
        labels_ref = Reference(ws_chart_data, min_col=1, min_row=2, max_row=21)
        values_ref = Reference(ws_chart_data, min_col=2, min_row=2, max_row=21)
        
        chart.add_data(values_ref, titles_from_data=False)
        chart.set_categories(labels_ref)
        ws_analysis.add_chart(chart, "B20")

    def _create_chart_data_sheet(self):
        if "ChartData" in self.wb.sheetnames:
            ws_chart_data = self.wb["ChartData"]
        else:
            ws_chart_data = self.wb.create_sheet("ChartData")
        ws_chart_data.sheet_state = "hidden"
        
        ws_chart_data['A1'] = "Dynamic Labels"
        ws_chart_data['B1'] = "Dynamic Values"
        
        MAX_BINS = 20
        start_col_static = 3
        
        for col_idx, metric in enumerate(self.metrics_list, 0):
            current_col = start_col_static + col_idx
            col_name = metric['Column Name']
            chart_data = metric.get('chart_data', {})
            
            ws_chart_data.cell(row=1, column=current_col, value=col_name)
            
            if chart_data:
                labels = chart_data.get('labels', [])
                values = chart_data.get('values', [])
                
                for i, label in enumerate(labels[:MAX_BINS]):
                    ws_chart_data.cell(row=2+i, column=current_col, value=label)
                for i, val in enumerate(values[:MAX_BINS]):
                    ws_chart_data.cell(row=22+i, column=current_col, value=val)
                    
        last_col_letter = get_column_letter(start_col_static + len(self.metrics_list) - 1)
        search_range_headers = f"$C$1:${last_col_letter}$1"
        search_range_labels = f"$C$2:${last_col_letter}$21"
        search_range_values = f"$C$22:${last_col_letter}$41"
        
        for i in range(MAX_BINS):
            row_idx = 2 + i
            match_part = f"MATCH('Column Analysis'!$C$5, ChartData!{search_range_headers}, 0)"
            label_formula = f"=INDEX(ChartData!{search_range_labels}, {i+1}, {match_part})"
            ws_chart_data.cell(row=row_idx, column=1, value=label_formula)
            value_formula = f"=INDEX(ChartData!{search_range_values}, {i+1}, {match_part})"
            ws_chart_data.cell(row=row_idx, column=2, value=value_formula)

    def _write_styled_table(self, ws, data, start_row, start_col, width_key=35, width_val=20, title_color="2F5597"):
        header_fill = PatternFill(start_color=title_color, end_color=title_color, fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        row_fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        border_color = "BFBFBF"
        thin_border = Border(left=Side(style='thin', color=border_color), 
                             right=Side(style='thin', color=border_color), 
                             top=Side(style='thin', color=border_color), 
                             bottom=Side(style='thin', color=border_color))
        
        header_row = start_row
        ws.merge_cells(start_row=header_row, start_column=start_col, end_row=header_row, end_column=start_col+1)
        cell_header = ws.cell(row=header_row, column=start_col, value=data[0][0])
        cell_header.font = header_font
        cell_header.fill = header_fill
        cell_header.alignment = Alignment(horizontal="center", vertical="center")
        cell_header.border = thin_border
        ws.cell(row=header_row, column=start_col+1).border = thin_border
        
        for i, row_data in enumerate(data[1:], 1):
            current_row = start_row + i
            cell_key = ws.cell(row=current_row, column=start_col, value=row_data[0])
            cell_key.font = Font(size=11, color="404040")
            cell_key.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell_key.border = thin_border
            
            val = row_data[1] if len(row_data) > 1 else ""
            cell_val = ws.cell(row=current_row, column=start_col + 1, value=val)
            cell_val.font = Font(bold=True, size=11, color="000000")
            cell_val.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            cell_val.border = thin_border
            
            fill = row_fill_even if i % 2 == 0 else row_fill_odd
            cell_key.fill = fill
            cell_val.fill = fill
            
        ws.column_dimensions[get_column_letter(start_col)].width = width_key
        ws.column_dimensions[get_column_letter(start_col + 1)].width = width_val

    def _create_styled_metric_table(self, ws, title, data, start_row, start_col, width_key=30, width_val=20):
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        row_fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        border_color = "BFBFBF"
        thin_border = Border(left=Side(style='thin', color=border_color), 
                             right=Side(style='thin', color=border_color), 
                             top=Side(style='thin', color=border_color), 
                             bottom=Side(style='thin', color=border_color))
        
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+1)
        cell_header = ws.cell(row=start_row, column=start_col, value=title)
        cell_header.font = header_font
        cell_header.fill = header_fill
        cell_header.alignment = Alignment(horizontal="center", vertical="center")
        cell_header.border = thin_border
        ws.cell(row=start_row, column=start_col+1).border = thin_border
        
        for i, (label, metric_name, fmt) in enumerate(data, 1):
            current_row = start_row + i
            cell_label = ws.cell(row=current_row, column=start_col, value=label)
            cell_label.font = Font(size=11, color="404040")
            cell_label.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell_label.border = thin_border
            
            formula = f"""=INDEX('Column Metrics'!$A:$ZZ, MATCH($C$5, 'Column Metrics'!$A:$A, 0), MATCH("{metric_name}", 'Column Metrics'!$1:$1, 0))"""
            cell_val = ws.cell(row=current_row, column=start_col+1, value=formula)
            cell_val.font = Font(bold=True, size=11, color="000000")
            cell_val.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            cell_val.number_format = fmt
            cell_val.border = thin_border
            
            fill = row_fill_even if i % 2 == 0 else row_fill_odd
            cell_label.fill = fill
            cell_val.fill = fill
            
        ws.column_dimensions[get_column_letter(start_col)].width = width_key
        ws.column_dimensions[get_column_letter(start_col + 1)].width = width_val

class NpEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if isinstance(obj, (datetime, da.Array)):
             return str(obj)
        return super(NpEncoder, self).default(obj)

class EDAOrchestrator:
    """Orchestrates the EDA process."""
    def __init__(self, args):
        self.args = args

    def run(self):
        start_time = datetime.now()
        print(f"Started execution at: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")

        # 1. Load Data
        dask_args = {}
        if self.args.dask_args:
            try:
                dask_args = json.loads(self.args.dask_args)
                print(f"Using custom Dask arguments: {dask_args}")
            except json.JSONDecodeError as e:
                print(f"Error parsing --dask-args: {e}")
                sys.exit(1)

        loader = DataLoader(Path(self.args.input_file), dask_args=dask_args)
        ddf = loader.load_data()
        
        # 2. Calculate Metrics
        calculator = MetricsCalculator(ddf)
        metrics_list, total_rows = calculator.calculate_metrics()
        
        # 3. Save Intermediate Metrics (Optional)
        if self.args.intermediate:
            print("Saving intermediate metrics to 'metrics.json'...")
            try:
                with open("metrics.json", "w") as f:
                    json.dump(metrics_list, f, cls=NpEncoder, indent=4)
                print("✓ Intermediate metrics saved.")
            except Exception as e:
                print(f"Warning: Could not save metrics.json: {e}")
        else:
            print("Skipping intermediate metrics dump (use --intermediate to enable).")
            
        # 4. Generate Report
        if self.args.output_file:
            output_filename = self.args.output_file
            if not output_filename.lower().endswith('.xlsx'):
                output_filename += '.xlsx'
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{Path(self.args.input_file).stem}_metrics_report_{timestamp}.xlsx"
            
        generator = ReportGenerator(metrics_list, total_rows, output_filename)
        generator.generate_report()

        end_time = datetime.now()
        duration = end_time - start_time
        
        print("\n" + "="*50)
        print("OPERATIONAL METRICS")
        print("="*50)
        print(f"Input File       : {self.args.input_file}")
        print(f"Output File      : {output_filename}")
        print(f"Total Rows       : {total_rows}")
        print(f"Total Columns    : {len(metrics_list)}")
        print(f"Start Time       : {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"End Time         : {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Total Duration   : {duration}")
        print("="*50 + "\n")

def main():
    # Sample Input Format:
    # python excel.py input.csv --output output.xlsx --intermediate
    args = parse_args()
    orchestrator = EDAOrchestrator(args)
    orchestrator.run()

if __name__ == "__main__":
    main()